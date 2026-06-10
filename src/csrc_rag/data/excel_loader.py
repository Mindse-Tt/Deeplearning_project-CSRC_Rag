from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from csrc_rag.data.schema import DATA_START_ROW, HEADER_ROW, ZH_HEADER_ROW, WorkbookSchema, normalize_text


def load_sheet(path: str | Path):
    workbook = load_workbook(path, read_only=False, data_only=True)
    return workbook[workbook.sheetnames[0]]


def read_schema(path: str | Path) -> WorkbookSchema:
    sheet = load_sheet(path)
    max_col = sheet.max_column
    english_headers = [normalize_text(sheet.cell(HEADER_ROW, col).value) or f"col_{col}" for col in range(1, max_col + 1)]
    chinese_headers = [normalize_text(sheet.cell(ZH_HEADER_ROW, col).value) or f"列{col}" for col in range(1, max_col + 1)]
    return WorkbookSchema(english_headers=english_headers, chinese_headers=chinese_headers)


def iter_records(path: str | Path) -> list[dict[str, Any]]:
    sheet = load_sheet(path)
    schema = read_schema(path)
    records: list[dict[str, Any]] = []

    for row_idx in range(DATA_START_ROW, sheet.max_row + 1):
        row: dict[str, Any] = {}
        empty_row = True
        for col_idx, header in enumerate(schema.english_headers, start=1):
            value = sheet.cell(row_idx, col_idx).value
            if value not in (None, ""):
                empty_row = False
            row[header] = value
        if empty_row:
            continue
        records.append(row)
    return records

